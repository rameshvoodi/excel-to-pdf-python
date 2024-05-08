from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, portrait, A0, A1, A2, A3, A4
from reportlab.lib.colors import Color, black
from reportlab.lib.units import inch
from openpyxl.styles import NamedStyle, Font
from openpyxl.cell.cell import Cell


def get_page_size(num_columns, num_rows):
    if num_columns <= 10 and num_rows <= 50:
        return A4
    elif num_columns <= 20 and num_rows <= 100:
        return A3
    elif num_columns <= 40 and num_rows <= 200:
        return A2
    elif num_columns <= 80 and num_rows <= 400:
        return A1
    elif num_columns > 80 or num_rows > 400:
        return A0
    else:
        return A4


def hex_to_rgb(hex_color):
    return tuple(int(hex_color[i : i + 2], 16) for i in (0, 2, 4))


def get_cell_value(cell):
    if cell.value is None:
        return ""
    elif cell.data_type == "n":
        return str(cell.value)
    elif cell.data_type == "s":
        return cell.value
    elif cell.data_type == "b":
        return "TRUE" if cell.value else "FALSE"
    elif cell.data_type == "d":
        return cell.value.strftime("%Y-%m-%d")
    elif cell.data_type == "e":
        return f"ERROR: {cell.value}"
    elif cell.data_type == "t":
        return cell.value.strftime("%H:%M:%S")
    else:
        return str(cell.value)


def excel_to_pdf(excel_file, pdf_file):
    workbook = load_workbook(excel_file, data_only=True)
    c = canvas.Canvas(pdf_file)

    for sheet_index, sheet in enumerate(workbook.worksheets):
        num_columns = len(sheet[1])
        num_rows = len(list(sheet.iter_rows()))
        page_size = get_page_size(num_columns, num_rows)
        c.setPageSize(page_size)

        if sheet_index > 0:
            c.showPage()

        page_width, page_height = page_size

        column_widths = [
            max(70, max(len(str(cell.value or "")) for cell in col) * 6) + 10
            for col in sheet.iter_cols()
        ]

        if len(column_widths) < num_columns:
            column_widths += [70] * (num_columns - len(column_widths))

        total_width = sum(column_widths)
        scale_factor = page_width / total_width
        column_widths = [width * scale_factor for width in column_widths]

        space_between_columns = [width * 0.1 for width in column_widths]
        column_widths = [
            width + space for width, space in zip(column_widths, space_between_columns)
        ]

        y = page_height
        rows = list(sheet.iter_rows())

        for index, row in enumerate(rows):
            if all(cell.value is None for cell in row):
                continue

            x = 0
            row_height = (
                max(sheet.row_dimensions[cell.row].height or 20 for cell in row)
                * 1.33333
            )
            if index == 0:
                row_height *= 2

            column_index = 0
            for cell in row:
                cell_width = column_widths[column_index]

                hex_color = None
                if cell.fill.bgColor:
                    color = cell.fill.bgColor
                    if color.auto:
                        pass
                    elif color.indexed is not None:
                        pass
                    elif color.type == "rgb":
                        hex_color = "{:02x}{:02x}{:02x}".format(
                            color.rgb.r, color.rgb.g, color.rgb.b
                        )

                if hex_color and len(hex_color) == 6:
                    rgb = hex_to_rgb(hex_color)
                    c.setFillColor(Color(rgb[0] / 255, rgb[1] / 255, rgb[2] / 255))
                    c.rect(x, y, cell_width, row_height, fill=1)

                text_object = c.beginText(x + 5, y - row_height + 10)
                c.setFillColor(black)

                text_object.setFont("Helvetica", cell.font.sz or 12)
                text_object.textLine(get_cell_value(cell))
                c.drawText(text_object)

                c.setStrokeColor(black)
                c.rect(x, y - row_height, cell_width, row_height)

                x += cell_width
                column_index += 1

            y -= row_height

    c.save()


excel_to_pdf("input.xlsx", "output.pdf")
